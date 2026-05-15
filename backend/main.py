import io
import json
import logging
import time
from urllib.parse import quote

from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook

from executor import execute_mapping
from headers import SYS_DATA_START_ROW, read_two_row_headers, read_user_table
from llm_parser import parse_rules


logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("excel-mapping")


app = FastAPI(title="Excel 字段映射填充工具")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173"],
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["X-Execute-Log", "Content-Disposition"],
)


def _validate_xlsx(upload: UploadFile, label: str) -> None:
    name = (upload.filename or "").lower()
    if not name.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail=f"{label} 必须是 .xlsx 文件")


async def _read_bytes(upload: UploadFile) -> bytes:
    data = await upload.read()
    await upload.seek(0)
    return data


@app.post("/api/sheets")
async def get_sheets(
    user_file: UploadFile = File(...),
    sys_file: UploadFile = File(...),
):
    _validate_xlsx(user_file, "用户表")
    _validate_xlsx(sys_file, "系统表")

    user_bytes = await _read_bytes(user_file)
    sys_bytes = await _read_bytes(sys_file)

    try:
        user_wb = load_workbook(io.BytesIO(user_bytes), read_only=True, data_only=True)
        sys_wb = load_workbook(io.BytesIO(sys_bytes), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"读取 Excel 失败: {e}")

    return {
        "user_sheets": user_wb.sheetnames,
        "sys_sheets": sys_wb.sheetnames,
    }


@app.post("/api/execute")
async def execute(
    user_file: UploadFile = File(...),
    sys_file: UploadFile = File(...),
    user_sheet: str = Form(...),
    sys_sheet: str = Form(...),
    rules: str = Form(...),
):
    _validate_xlsx(user_file, "用户表")
    _validate_xlsx(sys_file, "系统表")

    if not rules.strip():
        raise HTTPException(status_code=400, detail="规则文本不能为空")

    timings: dict[str, float] = {}
    t = time.time()

    user_bytes = await _read_bytes(user_file)
    sys_bytes = await _read_bytes(sys_file)
    timings["read_bytes"] = round(time.time() - t, 2); t = time.time()

    try:
        user_df, user_headers, user_letter_map = read_user_table(user_bytes, user_sheet)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"读取用户表失败: {e}")
    timings["read_user_table"] = round(time.time() - t, 2)
    log.info(f"用户表读取完成：{len(user_df)} 行 × {len(user_headers)} 列，耗时 {timings['read_user_table']}s")
    t = time.time()

    try:
        sys_wb = load_workbook(io.BytesIO(sys_bytes), keep_links=False)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"读取系统表失败: {e}")
    timings["load_sys_wb"] = round(time.time() - t, 2)
    log.info(f"系统表加载完成（openpyxl 全功能模式），耗时 {timings['load_sys_wb']}s")
    t = time.time()

    if sys_sheet not in sys_wb.sheetnames:
        raise HTTPException(status_code=400, detail=f"系统表中不存在 Sheet「{sys_sheet}」")
    sys_ws = sys_wb[sys_sheet]
    sys_headers = read_two_row_headers(sys_ws)
    timings["read_sys_headers"] = round(time.time() - t, 2); t = time.time()

    try:
        config = parse_rules(user_headers, sys_headers, user_letter_map, rules)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"规则解析失败: {e}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"调用 LLM 失败: {e}")
    timings["llm_parse"] = round(time.time() - t, 2)
    log.info(f"LLM 解析完成，耗时 {timings['llm_parse']}s")
    t = time.time()

    key_mapping = config.get("key_mapping") or {}
    if not key_mapping.get("user_col") or not key_mapping.get("sys_col"):
        raise HTTPException(status_code=400, detail="规则中未声明主键，请补充「主键：用户表[列] 对应 系统表[列]」")

    try:
        sys_wb, exec_log = execute_mapping(
            config,
            user_df,
            user_letter_map,
            sys_wb,
            sys_sheet,
            sys_data_start_row=SYS_DATA_START_ROW,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    timings["execute_mapping"] = round(time.time() - t, 2)
    log.info(
        f"映射执行完成：筛选 {exec_log['filtered_count']} 行，更新 {exec_log['updated_count']} 行，"
        f"跳过 {exec_log['skipped_count']} 行，耗时 {timings['execute_mapping']}s"
    )
    t = time.time()

    out = io.BytesIO()
    sys_wb.save(out)
    out.seek(0)
    timings["save_wb"] = round(time.time() - t, 2)
    log.info(f"系统表保存完成，耗时 {timings['save_wb']}s")

    exec_log["timings"] = timings
    log.info(f"全流程 timings: {timings}")

    log_json = json.dumps(exec_log, ensure_ascii=False, default=str)
    headers = {
        "X-Execute-Log": quote(log_json),
        "Content-Disposition": "attachment; filename*=UTF-8''" + quote("系统表_已填充.xlsx"),
    }
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
