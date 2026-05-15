"""
双层表头解析。

系统表：第 1 行 + 第 2 行表头，第 3 行计数（跳过），第 4 行起数据
用户表：第 1 行 + 第 2 行表头，第 3 行起数据

约定（两表通用）：
- 纯属性列：第 1 行有名字（与第 2 行垂直合并，所以第 2 行 value=None）
- 活动列：第 1 行是横向合并的活动名，第 2 行是子列名 → 扁平化为 "活动名.子列名"
- 用户表特殊情况：第 1 行整行为空，第 2 行直接是列名 → 扁平化为 "子列名"

性能策略：
- 用户表只读一次（pandas + python-calamine 引擎，Rust 流式实现）
- 系统表的全功能加载只发生在 main.py 里一次，配合 read_two_row_headers 读表头
"""
import io
from typing import Optional

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


SYS_DATA_START_ROW = 4
USER_DATA_START_ROW = 3


def _resolve_merged_value(ws: Worksheet, row: int, col: int):
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    merged_ranges = getattr(ws, "merged_cells", None)
    if merged_ranges is None:
        return None
    for mr in merged_ranges.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            anchor = ws.cell(mr.min_row, mr.min_col)
            return anchor.value
    return None


def _flat_name(top, bottom) -> str:
    top_s = "" if top is None else str(top).strip()
    bot_s = "" if bottom is None else str(bottom).strip()
    if top_s and bot_s and top_s != bot_s:
        return f"{top_s}.{bot_s}"
    return top_s or bot_s


def read_two_row_headers(ws: Worksheet) -> list[str]:
    """从已加载的 openpyxl Worksheet 读取双层表头，返回扁平后的列名列表。"""
    max_col = ws.max_column or 0
    result = []
    for col in range(1, max_col + 1):
        top = _resolve_merged_value(ws, 1, col)
        bot = ws.cell(2, col).value
        result.append(_flat_name(top, bot))
    return result


def _clean_cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    s = str(v).strip()
    if s.startswith("Unnamed:"):
        return ""
    return s


def _read_excel_fast(file_bytes: bytes, sheet_name: str, **kwargs) -> pd.DataFrame:
    """优先用 calamine 引擎读取；不可用时回落到 openpyxl 引擎。"""
    try:
        return pd.read_excel(
            io.BytesIO(file_bytes),
            sheet_name=sheet_name,
            engine="calamine",
            **kwargs,
        )
    except (ImportError, ValueError):
        return pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, **kwargs)


def read_user_table(file_bytes: bytes, sheet_name: str) -> tuple[pd.DataFrame, list[str], dict[str, str]]:
    """
    读取用户表（双层表头，第 3 行起数据）。返回 (DataFrame, headers, letter_map)。
    单次流式读取，依赖 pandas + calamine。
    合并单元格通过 pandas 的 MultiIndex + forward-fill 还原：
    - 横向合并的活动名只在锚点列有值，其余列为 NaN，forward-fill 把活动名"摊开"到所有子列
    - 垂直合并的属性列在第 2 行为 NaN，会被 _clean_cell 当作空字符串处理
    """
    df = _read_excel_fast(file_bytes, sheet_name, header=[0, 1])

    level0_raw = df.columns.get_level_values(0).tolist()
    level1_raw = df.columns.get_level_values(1).tolist()

    level0_filled = []
    last = ""
    for v in level0_raw:
        s = _clean_cell(v)
        if not s:
            level0_filled.append(last)
        else:
            last = s
            level0_filled.append(s)

    headers: list[str] = []
    for top, bot_raw in zip(level0_filled, level1_raw):
        bot = _clean_cell(bot_raw)
        if top and bot and top != bot:
            headers.append(f"{top}.{bot}")
        else:
            headers.append(top or bot)

    df.columns = headers
    letter_map = {get_column_letter(i + 1): h for i, h in enumerate(headers) if h}
    return df, headers, letter_map


def resolve_user_col_ref(ref, headers: list[str], letter_map: dict[str, str]) -> Optional[str]:
    """
    解析用户表中的列引用，返回对应的扁平表头名。无法解析返回 None。

    支持：
    - 直接的扁平表头名：'工号'、'活动A.计划开始'
    - 带"列"后缀的表头名：'工号列' → '工号'
    - Excel 列字母 + "列" 后缀：'P列' → letter_map['P']
    - 纯 Excel 列字母：'P' / 'AH' → letter_map[...]
    """
    if ref is None:
        return None
    s = str(ref).strip()
    if not s:
        return None
    candidates = [s]
    if s.endswith("列"):
        candidates.append(s[:-1])
    for cand in candidates:
        if cand in headers:
            return cand
        upper = cand.upper()
        if upper in letter_map:
            return letter_map[upper]
    return None
