from typing import Any

import pandas as pd
from openpyxl import Workbook

from compute import COMPUTE_FUNCTIONS
from headers import read_two_row_headers, resolve_user_col_ref


def _apply_filters(
    user_df: pd.DataFrame,
    filters: list[dict],
    user_letter_map: dict[str, str],
    warnings: list[str],
) -> pd.DataFrame:
    df = user_df.copy()
    for f in filters or []:
        raw_col = f.get("column")
        op = f.get("operator")
        val = f.get("value")
        col = resolve_user_col_ref(raw_col, df.columns.tolist(), user_letter_map)
        if col is None:
            warnings.append(f"筛选跳过：用户表中找不到列「{raw_col}」")
            continue
        col_data = df[col].astype(str)
        target = "" if val is None else str(val)
        if op == "equals":
            df = df[col_data == target]
        elif op == "contains":
            df = df[col_data.str.contains(target, na=False, regex=False)]
        elif op == "not_equals":
            df = df[col_data != target]
        else:
            warnings.append(f"筛选跳过：未知运算符「{op}」")
    return df


def _set_cell(ws, row_num: int, headers: list[str], col_name: str, value: Any, warnings: list[str]) -> bool:
    if col_name is None or col_name not in headers:
        warnings.append(f"系统表中不存在列「{col_name}」，跳过")
        return False
    col_idx = headers.index(col_name)
    if isinstance(value, float) and pd.isna(value):
        value = ""
    ws.cell(row=row_num, column=col_idx + 1).value = value
    return True


def _get_user_val(user_row, raw_col, df_columns, letter_map, warnings, label):
    col = resolve_user_col_ref(raw_col, df_columns, letter_map)
    if col is None:
        warnings.append(f"{label} 跳过：用户表中找不到列「{raw_col}」")
        return None, None
    return col, user_row.get(col)


def execute_mapping(
    config: dict,
    user_df: pd.DataFrame,
    user_letter_map: dict[str, str],
    sys_wb: Workbook,
    sys_sheet: str,
    sys_data_start_row: int = 4,
) -> tuple[Workbook, dict]:
    warnings: list[str] = []

    if sys_sheet not in sys_wb.sheetnames:
        raise ValueError(f"系统表中不存在 Sheet「{sys_sheet}」")
    ws = sys_wb[sys_sheet]

    headers = read_two_row_headers(ws)
    df_columns = user_df.columns.tolist()

    key_mapping = config.get("key_mapping") or {}
    raw_user_key = key_mapping.get("user_col")
    raw_sys_key = key_mapping.get("sys_col")
    if not raw_user_key or not raw_sys_key:
        raise ValueError("缺少主键声明 (key_mapping)")

    user_key_col = resolve_user_col_ref(raw_user_key, df_columns, user_letter_map)
    if user_key_col is None:
        raise ValueError(f"用户表中找不到主键列「{raw_user_key}」")
    if raw_sys_key not in headers:
        raise ValueError(f"系统表中找不到主键列「{raw_sys_key}」")

    sys_key_idx = headers.index(raw_sys_key)

    sys_index: dict[str, int] = {}
    for row_num, row in enumerate(ws.iter_rows(min_row=sys_data_start_row), start=sys_data_start_row):
        if sys_key_idx >= len(row):
            continue
        key_val = row[sys_key_idx].value
        if key_val is None:
            continue
        key_str = str(key_val).strip()
        if key_str:
            sys_index[key_str] = row_num

    filtered_df = _apply_filters(user_df, config.get("filters") or [], user_letter_map, warnings)
    filtered_count = len(filtered_df)

    updated_count = 0
    skipped_keys: list[str] = []
    mappings = config.get("mappings") or []

    for _, user_row in filtered_df.iterrows():
        raw_key = user_row.get(user_key_col)
        if raw_key is None or (isinstance(raw_key, float) and pd.isna(raw_key)):
            skipped_keys.append("")
            continue
        key_val = str(raw_key).strip()
        row_num = sys_index.get(key_val)
        if row_num is None:
            skipped_keys.append(key_val)
            continue

        for mapping in mappings:
            mtype = mapping.get("type")

            if mtype == "copy":
                _, value = _get_user_val(
                    user_row, mapping.get("source_column"), df_columns, user_letter_map, warnings, "copy"
                )
                if value is None and resolve_user_col_ref(
                    mapping.get("source_column"), df_columns, user_letter_map
                ) is None:
                    continue
                _set_cell(ws, row_num, headers, mapping.get("target_column"), value, warnings)

            elif mtype == "conditional":
                src_col, src_val_raw = _get_user_val(
                    user_row, mapping.get("source_column"), df_columns, user_letter_map, warnings, "conditional"
                )
                if src_col is None:
                    continue
                src_val = "" if src_val_raw is None or (isinstance(src_val_raw, float) and pd.isna(src_val_raw)) else str(src_val_raw)
                matched = False
                for cond in mapping.get("conditions") or []:
                    op = cond.get("operator")
                    cval = cond.get("value")
                    hit = False
                    if op == "equals":
                        hit = src_val == ("" if cval is None else str(cval))
                    elif op == "contains":
                        hit = ("" if cval is None else str(cval)) in src_val
                    elif op == "not_equals":
                        hit = src_val != ("" if cval is None else str(cval))
                    elif op == "in":
                        cand = cval if isinstance(cval, list) else [cval]
                        hit = src_val in [("" if c is None else str(c)) for c in cand]
                    if hit:
                        _set_cell(ws, row_num, headers, cond.get("target_column"), cond.get("target_value"), warnings)
                        matched = True
                        break
                if not matched:
                    default = mapping.get("default")
                    if default:
                        _set_cell(ws, row_num, headers, default.get("target_column"), default.get("target_value"), warnings)

            elif mtype == "compute":
                fn_name = mapping.get("function")
                fn = COMPUTE_FUNCTIONS.get(fn_name)
                if fn is None:
                    warnings.append(f"compute 跳过：未知函数「{fn_name}」")
                    continue
                raw_src_cols = mapping.get("source_columns") or []
                resolved = [resolve_user_col_ref(c, df_columns, user_letter_map) for c in raw_src_cols]
                missing = [raw for raw, r in zip(raw_src_cols, resolved) if r is None]
                if missing:
                    warnings.append(f"compute 跳过：用户表中找不到列 {missing}")
                    continue
                result = fn(user_row, resolved)
                _set_cell(ws, row_num, headers, mapping.get("target_column"), result, warnings)

            else:
                warnings.append(f"未知 mapping 类型：{mtype}")

        updated_count += 1

    log = {
        "filtered_count": filtered_count,
        "updated_count": updated_count,
        "skipped_count": len(skipped_keys),
        "skipped_keys": skipped_keys,
        "warnings": warnings,
        "config": config,
    }
    return sys_wb, log
